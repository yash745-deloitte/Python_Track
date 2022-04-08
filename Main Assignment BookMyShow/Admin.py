from openpyxl import *

class Admin:
    def admin(self, id, pas):
        path = "resources\\Book1.xlsx"
        wb = load_workbook(path)
        sheet = wb["Sheet3"]
        rows = sheet.max_row
        print('''******Welcome Admin******* 
1. Add New Movie Info 
2. Edit Movie Info 
3. Delete Movies 
4.Logout 
        ''')
        print(" ")
        b = int(input("Enter your choice: "))
        if b==1:
            path = "resources\\Book1.xlsx"
            wb = load_workbook(path)
            sheet = wb['Sheet3']
            rows = sheet.max_row
            print("*****Create New Movie*****")
            sheet.cell(row=rows + 1, column=1).value = input("Title : ")
            sheet.cell(row=rows + 1, column=2).value = input("Genre : ")
            sheet.cell(row=rows + 1, column=3).value = input("Length : ")
            sheet.cell(row=rows + 1, column=4).value = input("Cast : ")
            sheet.cell(row=rows + 1, column=5).value = input("Director : ")
            sheet.cell(row=rows + 1, column=6).value = input("Admin-Rating : ")
            sheet.cell(row=rows + 1, column=7).value = input("User-Rating : ")
            sheet.cell(row=rows + 1, column=8).value = input("Timing of Show 1 : ")
            sheet.cell(row=rows + 1, column=10).value = input("Timing of Show 2 : ")
            sheet.cell(row=rows + 1, column=12).value = input("Timing of Show 3 : ")
            sheet.cell(row=rows, column=9).value = int(input("1st Show Seats : "))
            sheet.cell(row=rows, column=11).value = int(input("2nd Show Seats : "))
            sheet.cell(row=rows, column=13).value = int(input("3rd Show Seats : "))
            # sheet.cell(row=rows + 1, column=9).value =  sheet.cell(row=rows + 1, column=14).value
            # sheet.cell(row=rows + 1, column=11).value = sheet.cell(row=rows + 1, column=14).value
            # sheet.cell(row=rows + 1, column=13).value = sheet.cell(row=rows + 1, column=14).value
            wb.save("resources\\Book1.xlsx")
            print("Movie Sucessfully Added")
            return
        elif b==2:
            path = "resources\\Book1.xlsx"
            wb = load_workbook(path)
            sheet = wb['Sheet3']
            rows = sheet.max_row
            while True:
                for i in range(2, rows + 1):
                    print(i - 1, sheet.cell(row=i, column=1).value)
                x = int(input("Enter Movie Number : "))
                x = x + 1
                if x <= rows and x > 0:
                    rows = x
                    break
                else:
                    print("Invalid Input")

            print("*****Edit Movie*****")
            print(
                '''
1.Title
2.Genre
3.Length
4.Cast
5.Director
6.Admin-Rating
7.User-Rating
8.Timing of Show 1
9.Timing of Show 2
10.Timing of Show 3
11.Total Seats
                '''
            )
            print(" ")
            b = int(input("Enter your choice: "))
            if b == 1:
                sheet.cell(row=rows, column=1).value = input("Title : ")
            elif b == 2:
                sheet.cell(row=rows, column=2).value = input("Genre : ")
            elif b == 3:
                sheet.cell(row=rows, column=3).value = input("Length : ")
            elif b == 4:
                sheet.cell(row=rows, column=4).value = input("Cast : ")
            elif b == 5:
                sheet.cell(row=rows, column=5).value = input("Director : ")
            elif b == 6:
                sheet.cell(row=rows, column=6).value = input("Admin-Rating : ")
            elif b == 7:
                sheet.cell(row=rows, column=7).value = input("User-Rating : ")
            elif b == 8:
                sheet.cell(row=rows, column=8).value = input("Timing of Show 1 : ")
            elif b == 9:
                sheet.cell(row=rows, column=10).value = input("Timing of Show 2 : ")
            elif b == 10:
                sheet.cell(row=rows, column=12).value = input("Timing of Show 3 : ")
            elif b == 11:
                sheet.cell(row=rows, column=9).value = int(input("1st Show Seats : "))  # sheet.cell(row=rows + 1, column=14).value
                sheet.cell(row=rows, column=11).value = int(input("2nd Show Seats : "))  # sheet.cell(row=rows + 1, column=14).value
                sheet.cell(row=rows, column=13).value = int(input("3rd Show Seats : "))  # shee, column=14).value
            wb.save("resources\\Book1.xlsx")
            print("Movie Sucessfully Edited")
        elif b == 3:
            path = "resources\\Book1.xlsx"
            wb = load_workbook(path)
            sheet = wb['Sheet3']
            rows = sheet.max_row
            while True:
                for i in range(2, rows + 1):
                    print(i - 1, sheet.cell(row=i, column=1).value)
                x = int(input("Enter Movie Number : "))
                if x <= rows and x > 0:
                    rows = x + 1
                    break
                else:
                    print("Invalid Input")
            sheet.delete_rows(rows)
            wb.save("resources\\Book1.xlsx")
            print("Movie Sucessfully Deleted")
        elif b == 4:
            return
        else:
            print("Invalid Input")

loginAdmin = Admin()