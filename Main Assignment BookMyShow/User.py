from openpyxl import *

class Userlogin:
    def user(self, id, pas):
        path = "resources\\Book1.xlsx"
        wb = load_workbook(path)
        sheet = wb["Sheet3"]
        rows = sheet.max_row

        mysets = 0
        while True:
            print("*****Welcome*****")
            for i in range(2, rows + 1):
                print(i-1,".", sheet.cell(row=i, column=1).value)

            print(rows,". Exit")

            mov = int(input("Enter Movie Number: "))
            if mov < rows and mov >= 1:
                print("*****Welcome User*****")
                print('''
1. Book Tickets 
2. Cancel Tickets 
3. User Rating
4. Go Back 
                                ''')
                ch = int(input("Enter Choice: "))
                if ch == 1:
                    print("*****Welcome User*****")
                    print("1.", sheet.cell(row=mov + 1, column=8).value)
                    print("2.", sheet.cell(row=mov + 1, column=10).value)
                    print("3.", sheet.cell(row=mov + 1, column=12).value)
                    d = int(input("Select show timings:"))
                    if d == 1:
                        seat = 9
                        timming = 8
                    elif d == 2:
                        seat = 11
                        timming = 10
                    elif d == 3:
                        seat = 13
                        timming = 12
                    else:
                        print("Invalid Input")
                        continue
                    print("Timing: ", sheet.cell(row=mov + 1, column=timming).value)
                    print("Remaining Seats: ", sheet.cell(row=mov + 1, column=seat).value)

                    bookseat = int(input("Enter number of seats: "))
                    if sheet.cell(row=mov + 1, column=seat).value >= bookseat:
                        mysets = mysets + bookseat
                        sheet.cell(row=mov + 1, column=seat).value -= bookseat
                        print(mysets, "Total seats booked...")
                        wb.save(path)
                    else:
                        print("Invalid number of seats to book")
                elif ch == 2:
                    print("*****Welcome User*****")
                    cancelseat = int(input("No. of seats you want to cancel: "))
                    if cancelseat <= mysets:
                        mysets = mysets - cancelseat
                    else:
                        print("Invalid number of seats to cancel")

                elif ch == 4:
                    pass
                elif ch == 3:
                    print("*****Welcome User*****")
                    # user_rating = int(input("Your Rating for this movie: "))
                    sheet.cell(row=rows, column=7).value = input("Your Rating for this movie: ")
                    wb.save(path)

            elif mov == rows:
                break
            else:
                print("Invalid Input")
loginUser = Userlogin()