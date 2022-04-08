from openpyxl import *
from User import *
from Admin import *

class verify:
    def check(self, id, pas, sheet):
        path = "resources\\Book1.xlsx"
        wb = load_workbook(path)
        sheet = wb[sheet]
        rows = sheet.max_row

        for i in range(2, rows + 1):
            idobj = sheet.cell(row=i, column=1)
            if idobj.value == id:
                pasobj = sheet.cell(row=i, column=2)
                if pasobj.value == pas:
                    return True
        return False
class Login:
    def log(self):
        print("*****Welcome To Book My Show****")
        print("1. Admin Login")
        print("2. User Login")
        print("3. Back to Main Menu")
        print(" ")
        b = int(input("Enter your choice: "))

        if b == 1:
            print("*****Hello Admin, Pls Login*****")
            print(" ")
            id = input("E-mail: ")
            pas = input("Password: ")
            veri = verify()
            if veri.check(id, pas, "Sheet2"):
                loginUser = Admin()
                loginUser.admin(id, pas)
            else:
                print("INVALID INPUT")

        elif b == 2:
            print("*****Hello User, Pls Login*****")
            print(" ")
            id = input("E-mail: ")
            pas = input("Password: ")
            veri = verify()
            if veri.check(id, pas, "Sheet1"):
                loginUser = Userlogin()
                loginUser.user(id, pas)
            else:
                print("INVALID INPUT")

        elif b == 3:
            return
        else:
            print("Invalid Input")

loginObj = Login()
del loginObj

class Register:
    def newuser(self):
        path = "resources\\Book1.xlsx"
        wb = load_workbook(path)
        sheet = wb['Sheet1']
        rows = sheet.max_row
        print("*****Create New Account*****")
        sheet.cell(row = rows+1, column = 3).value = input("Name : ")
        sheet.cell(row = rows+1, column = 1).value = input("Email : ")
        sheet.cell(row = rows+1, column = 4).value = input("Phone : ")
        sheet.cell(row = rows+1, column = 5).value = input("Age : ")
        sheet.cell(row = rows+1, column = 2).value = input("Password : ")
        wb.save("resources\\Book1.xlsx")
        print("User Sucessfully Registered")

registerObj = Register()
del registerObj